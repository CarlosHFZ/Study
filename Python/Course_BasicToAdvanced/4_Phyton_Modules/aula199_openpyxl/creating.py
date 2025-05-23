# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Nome para a planilha
sheet_name = 'My Sheet'

workbook = Workbook()

# criamos a planilha
workbook.create_sheet(sheet_name, 0)

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover a planilha
workbook.remove(workbook['Sheet'])

print(workbook.sheetnames)

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'Score')

students = [
    # name      age score
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_collum in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_collum)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
